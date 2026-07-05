"""
Views for Analytics API
Provides descriptive analytics and data export, plus response-time statistics for the dashboard.
"""
from datetime import datetime
from io import BytesIO

from django.http import HttpResponse
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response

from app.core.logger import iot_logger
from app.core.mongodb_client import mongodb_service

try:
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


@api_view(['GET'])
def export_data(request):
    """
    Export sensor data to Excel or CSV.

    Query parameters:
    - measurement: Sensor measurement type (optional)
    - device_id: Device ID (optional)
    - start_time: ISO start time (optional)
    - end_time: ISO end time (optional)
    - format: 'xlsx' or 'csv' (default: xlsx)
    - limit: max rows (default 10000)
    - flatten_tags: '1'|'0' (default: 1) -> converts tags dict into columns tag_<k>
    """
    if not PANDAS_AVAILABLE:
        return Response({"status": "error", "message": "Pandas + openpyxl required for export"},
                        status=status.HTTP_503_SERVICE_UNAVAILABLE)

    try:
        measurement = request.query_params.get('measurement')
        device_id = request.query_params.get('device_id')
        fmt = request.query_params.get('format', 'xlsx').lower()
        limit = int(request.query_params.get('limit', 10000))
        flatten_tags = request.query_params.get('flatten_tags', '1') != '0'

        start_time = request.query_params.get('start_time')
        end_time = request.query_params.get('end_time')

        if measurement:
            data = mongodb_service.query_sensor_data(
                measurement=measurement,
                device_id=device_id,
                limit=limit,
                start_time=start_time,
                stop_time=end_time,
                default_time_window=False
            )
        else:
            measurements = mongodb_service.get_distinct_measurements() or [
                "heart_rate", "blood_pressure_systolic", "blood_pressure_diastolic",
                "body_temperature", "oxygen_saturation", "glucose_level", "activity_steps"
            ]
            all_data = []
            per = max(1, limit // max(1, len(measurements)))
            for meas in measurements[:10]:
                all_data.extend(mongodb_service.query_sensor_data(
                    measurement=meas,
                    device_id=device_id,
                    limit=per,
                    start_time=start_time,
                    stop_time=end_time,
                    default_time_window=False
                ))
            data = all_data

        if not data:
            return Response({"status": "error", "message": "No data found for export"},
                            status=status.HTTP_404_NOT_FOUND)

        df = pd.DataFrame(data)
        if 'time' in df.columns:
            df = df.rename(columns={'time': 'timestamp'})

        # Make export "analysis-ready": flatten tags into columns tag_<k>
        if 'tags' in df.columns and flatten_tags:
            tags_norm = df['tags'].apply(lambda x: x if isinstance(x, dict) else {})
            tag_df = pd.json_normalize(tags_norm).add_prefix("tag_")
            df = pd.concat([df.drop(columns=['tags']), tag_df], axis=1)
        elif 'tags' in df.columns:
            df['tags'] = df['tags'].apply(lambda x: str(x) if x is not None else "")

        if fmt == 'csv':
            output = BytesIO()
            df.to_csv(output, index=False)
            output.seek(0)
            resp = HttpResponse(output.read(), content_type='text/csv')
            resp[
                'Content-Disposition'] = f'attachment; filename="health_monitoring_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
            return resp

        if fmt != 'xlsx':
            return Response({"status": "error", "message": "Unsupported format. Use xlsx or csv."},
                            status=status.HTTP_400_BAD_REQUEST)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Sensor Data', index=False)
            ws = writer.sheets['Sensor Data']

            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for column in ws.columns:
                max_length = 0
                col_letter = column[0].column_letter
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

        output.seek(0)
        resp = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp[
            'Content-Disposition'] = f'attachment; filename="health_monitoring_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return resp

    except Exception as e:
        iot_logger.error(f"Export error: {e}", source="analytics")
        return Response({"status": "error", "message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def analytics_summary(request):
    """
    Descriptive statistics for a measurement.

    Query parameters:
    - measurement: required
    - device_id: optional
    - window: '1h'|'24h'|'7d'|'30d' (default 24h) [used for aggregated plot preview]
    """
    import statistics

    try:
        measurement = request.query_params.get('measurement')
        if not measurement:
            return Response({"status": "error", "message": "measurement parameter is required"},
                            status=status.HTTP_400_BAD_REQUEST)

        device_id = request.query_params.get('device_id')
        window = request.query_params.get('window', '24h')

        aggregated = mongodb_service.get_aggregated_data(measurement=measurement, device_id=device_id, window=window,
                                                         aggregate='mean')
        historical = mongodb_service.query_sensor_data(measurement=measurement, device_id=device_id, limit=2000)

        values = [float(x.get('value')) for x in historical if x.get('value') is not None]
        if not values:
            return Response({"status": "error", "message": "No data found"}, status=status.HTTP_404_NOT_FOUND)

        summary = {
            "measurement": measurement,
            "device_id": device_id or "all",
            "window": window,
            "statistics": {
                "count": len(values),
                "mean": float(statistics.mean(values)),
                "median": float(statistics.median(values)),
                "std_dev": float(statistics.stdev(values)) if len(values) > 1 else 0.0,
                "min": float(min(values)),
                "max": float(max(values)),
                "range": float(max(values) - min(values)),
            },
            "aggregated_data": aggregated[:10] if aggregated else [],
            "timestamp": datetime.now().isoformat(),
        }

        if len(values) >= 10:
            s = sorted(values)
            n = len(s)
            summary["statistics"]["p25"] = float(s[int(n * 0.25)])
            summary["statistics"]["p75"] = float(s[int(n * 0.75)])
            summary["statistics"]["p95"] = float(s[min(n - 1, int(n * 0.95))])

        return Response({"status": "success", "summary": summary}, status=status.HTTP_200_OK)

    except Exception as e:
        iot_logger.error(f"Analytics summary error: {e}", source="analytics")
        return Response({"status": "error", "message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def response_time_report(request):
    """
    Response-time statistics from simulation cycles.

    Query parameters:
    - limit: number of recent cycles (default 100)
    - start_time: ISO start (optional)
    - end_time: ISO end (optional)

    Response:
    {
      status, cycles_analyzed,
      statistics: { total: {avg,p95,max,...}, sensor_generation: {...}, ... }
    }
    """
    import statistics

    try:
        limit = int(request.query_params.get('limit', 100))

        start_time = request.query_params.get('start_time')
        end_time = request.query_params.get('end_time')
        start_dt = datetime.fromisoformat(start_time.replace('Z', '+00:00')) if start_time else None
        end_dt = datetime.fromisoformat(end_time.replace('Z', '+00:00')) if end_time else None

        rows = mongodb_service.query_response_times(limit=limit, start_time=start_dt, end_time=end_dt)

        if not rows:
            return Response({
                "status": "info",
                "message": "No response time data found. Run simulation cycles to generate data.",
                "cycles_analyzed": 0,
                "statistics": {}
            }, status=status.HTTP_200_OK)

        stages = ["sensor_generation", "edge_processing", "storage", "ml_prediction", "actuator_decision", "total"]
        buckets = {s: [] for s in stages}

        for r in rows:
            for s in stages:
                if r.get(s) is not None:
                    buckets[s].append(float(r[s]))

        def stats_for(values):
            if not values:
                return None
            svals = sorted(values)
            n = len(svals)
            return {
                "count": n,
                "avg": round(float(statistics.mean(values)), 4),
                "min": round(float(min(values)), 4),
                "max": round(float(max(values)), 4),
                "median": round(float(statistics.median(values)), 4),
                "p95": round(float(svals[min(n - 1, int(n * 0.95))]), 4),
            }

        out = {}
        for s in stages:
            st = stats_for(buckets[s])
            if st:
                out[s] = st

        return Response({
            "status": "success",
            "cycles_analyzed": len(rows),
            "time_range": {
                "start": start_dt.isoformat() if start_dt else None,
                "end": end_dt.isoformat() if end_dt else None
            },
            "statistics": out,
            "timestamp": datetime.now().isoformat()
        }, status=status.HTTP_200_OK)

    except Exception as e:
        iot_logger.error(f"Response time report error: {e}", source="analytics")
        return Response({"status": "error", "message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
